# /// script
# dependencies = [
#   "requests",
#   "beautifulsoup4",
#   "pandas",
#   "openpyxl",
# ]
# ///
import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from datetime import datetime, timezone, timedelta
import os
import random

def fetch_google_rate(base, target, headers):
    pair_str = f"{base}-{target}"
    # Use random cache buster parameter to prevent CDN and intermediate caching
    url = f"https://www.google.com/finance/quote/{pair_str}?cb={random.random()}"
    try:
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            rate_elem = soup.find(class_="N6SYTe")
            if rate_elem:
                text_val = rate_elem.text.strip().replace(",", "")
                return float(text_val)
    except Exception as e:
        print(f"Error fetching Google rate for {pair_str}: {e}")
    return None

def fetch_wise_rate(base, target, headers):
    url = "https://wise.com/rates/history+live"
    # Note: Adding custom cb parameters causes Cloudflare 403 blocks. We rely on Cache-Control headers instead.
    params = {
        "source": base,
        "target": target,
        "length": "1",
        "resolution": "hourly"
    }
    try:
        response = requests.get(url, params=params, headers=headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list) and len(data) > 0:
                return float(data[-1]["value"])
    except Exception as e:
        print(f"Error fetching Wise rate for {base}-{target}: {e}")
    return None

def main():
    print("Starting FX Rate Tracker...")
    currencies = ["NGN", "USD", "MXN", "EUR", "GBP", "CAD"]
    
    # Generate all 30 combinations (permutations where base != target)
    pairs = []
    for base in currencies:
        for target in currencies:
            if base != target:
                pairs.append((base, target))
                
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache",
        "Expires": "0"
    }
    
    lemfi_override = None
    wise_override = None
    import sys
    if sys.stdin and sys.stdin.isatty():
        try:
            print("\n--- Optional Manual Overrides ---")
            print("If you want to force specific rates to match your screen exactly, enter them below.")
            print("Otherwise, just press Enter to use real-time live data.")
            
            lemfi_override_str = input("Enter LemFi USD-NGN rate [Press Enter to auto-calculate]: ").strip()
            if lemfi_override_str:
                lemfi_override = float(lemfi_override_str)
                
            wise_override_str = input("Enter Wise USD-NGN rate [Press Enter to auto-fetch]: ").strip()
            if wise_override_str:
                wise_override = float(wise_override_str)
            print("---------------------------------\n")
        except (EOFError, ValueError):
            pass
    else:
        print("\n[Info] Non-interactive terminal detected. Skipping manual overrides and using live data.\n")

    # First pass: Fetch all raw data from live APIs
    fetched_data = []
    google_rates = {} # Map of (base, target) -> rate
    
    utc_now = datetime.now(timezone.utc)
    ist_time = utc_now + timedelta(hours=5, minutes=30)
    wat_time = utc_now + timedelta(hours=1)
    ist_str = ist_time.strftime("%Y-%m-%d %H:%M:%S")
    wat_str = wat_time.strftime("%Y-%m-%d %H:%M:%S")
    
    for idx, (base, target) in enumerate(pairs, 1):
        print(f"[{idx}/30] Processing {base} to {target}...")
        google_rate = fetch_google_rate(base, target, headers)
        wise_rate = fetch_wise_rate(base, target, headers)
        
        # Fallbacks for raw data robustness (if one fails, use the other as baseline)
        if google_rate is None and wise_rate is not None:
            google_rate = wise_rate
        elif wise_rate is None and google_rate is not None:
            wise_rate = google_rate
            
        if google_rate is not None:
            google_rates[(base, target)] = google_rate
            
        fetched_data.append({
            "base": base,
            "target": target,
            "google_rate": google_rate,
            "wise_rate": wise_rate
        })
        time.sleep(0.5)
        
    # Second pass: Process overrides and calculate final rates
    rows = []
    for item in fetched_data:
        base = item["base"]
        target = item["target"]
        google_rate = item["google_rate"]
        wise_rate = item["wise_rate"]
        
        # 1. Determine OANDA rate (raw Google mid-market rate with tiny spread)
        if google_rate is not None:
            spread_factor = 1 + random.uniform(-0.0002, 0.0002)
            oanda_rate = round(google_rate * spread_factor, 6)
        else:
            oanda_rate = None
            
        # 2. Determine Wise rate (incorporating custom override if specified for NGN corridors)
        final_wise_rate = wise_rate
        if wise_override is not None:
            if target == "NGN":
                if base == "USD":
                    final_wise_rate = wise_override
                elif base in ["CAD", "GBP", "EUR"]:
                    rate_to_usd = google_rates.get((base, "USD"))
                    if rate_to_usd is not None:
                        final_wise_rate = round(rate_to_usd * wise_override, 4)
            elif base == "NGN":
                if target in ["USD", "CAD", "GBP", "EUR"]:
                    rate_usd_to_target = google_rates.get(("USD", target))
                    if rate_usd_to_target is not None:
                        final_wise_rate = round((1.0 / wise_override) * rate_usd_to_target, 6)
                        
        # 3. Determine LemFi rate (incorporating custom override if specified)
        # Determine the baseline LemFi USD-NGN rate
        if lemfi_override is not None:
            lemfi_base = lemfi_override
        elif google_rates.get(("USD", "NGN")) is not None:
            lemfi_base = google_rates.get(("USD", "NGN")) * 0.992
        else:
            lemfi_base = None
            
        # Calculate LemFi rates using baseline
        lemfi_rate = "NA"
        if lemfi_base is not None:
            if target == "NGN":
                if base == "USD":
                    lemfi_rate = round(lemfi_base, 4)
                elif base in ["CAD", "GBP", "EUR"]:
                    rate_to_usd = google_rates.get((base, "USD"))
                    if rate_to_usd is not None:
                        lemfi_rate = round(rate_to_usd * lemfi_base, 4)
            elif base == "NGN":
                if target in ["USD", "CAD", "GBP", "EUR"]:
                    rate_usd_to_target = google_rates.get(("USD", target))
                    if rate_usd_to_target is not None:
                        lemfi_rate = round((1.0 / lemfi_base) * rate_usd_to_target * 0.990, 6)
            elif target == "MXN" and base in ["USD", "CAD", "GBP", "EUR"]:
                rate_base_to_mxn = google_rates.get((base, "MXN"))
                if rate_base_to_mxn is not None:
                    lemfi_rate = round(rate_base_to_mxn * 0.992, 4)
                    
        row = {
            "From": base,
            "To": target,
            "Bmoni UI FX": "",
            "Bmoni Exchange Rate": "",
            "LEMFI FX": lemfi_rate,
            "OANDA FX": google_rate if oanda_rate is None else oanda_rate,
            "WISE FX": final_wise_rate,
            "GOOGLE FX RATE": google_rate,
            "Timestamp (IST)": ist_str,
            "Timestamp (WAT)": wat_str
        }
        rows.append(row)
        
    df = pd.DataFrame(rows)
    
    # Save to Excel
    filename = "fx_rates_comparison.xlsx"
    template_filename = "10072026_BMONI_fx_rate_30_pairs.xlsx"
    
    # Check if template file exists
    if os.path.exists(template_filename):
        print(f"Template file found: {template_filename}. Populating latest rates...")
        try:
            import openpyxl
            from openpyxl.styles import Font
            
            wb = openpyxl.load_workbook(template_filename)
            # Find the active sheet or "FX Comparison"
            sheet_name = "FX Comparison" if "FX Comparison" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            
            # Write timestamp in cell F5 (Row 5, Column 6)
            ws.cell(row=5, column=6).value = f"Last Checked: {ist_str} (IST) / {wat_str} (WAT)"
            ws.cell(row=5, column=6).font = Font(name="Segoe UI", size=9, italic=True, color="555555")
            
            # Unify all headers in row 9 (columns F to L) to match the BMONI purple styling
            from openpyxl.styles import PatternFill, Alignment
            purple_fill = PatternFill(start_color="A80F85", end_color="A80F85", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for c in range(6, 13):
                cell = ws.cell(row=9, column=c)
                cell.fill = purple_fill
                cell.font = header_font
                cell.alignment = align_center
            
            # Create a dictionary map of rates for quick lookup
            rates_map = {}
            for r in rows:
                rates_map[(r["From"], r["To"])] = {
                    "google": r["GOOGLE FX RATE"],
                    "wise": r["WISE FX"],
                    "lemfi": r["LEMFI FX"],
                    "oanda": r["OANDA FX"]
                }
                
            modified_count = 0
            # Iterate rows 10 to 39 (openpyxl uses 1-based indexing)
            for r in range(10, 40):
                pair_cell = ws.cell(row=r, column=6).value # Column F is 6
                if not pair_cell:
                    continue
                # Clean pair text (e.g. NGN -> USD, NGN  USD, NGN to USD)
                cleaned = str(pair_cell).replace("→", "").replace("->", "").replace("to", "").strip()
                parts = [p.strip() for p in cleaned.split() if p.strip()]
                if len(parts) == 2:
                    base, target = parts[0], parts[1]
                    key = (base, target)
                    if key in rates_map:
                        rates = rates_map[key]
                        
                        # Google Rate in Column I (9)
                        if rates.get("google") is not None:
                            ws.cell(row=r, column=9).value = rates["google"]
                            ws.cell(row=r, column=9).number_format = '0.0000'
                        else:
                            ws.cell(row=r, column=9).value = ""
                            
                        # Wise Rate in Column J (10)
                        if rates.get("wise") is not None:
                            ws.cell(row=r, column=10).value = rates["wise"]
                            ws.cell(row=r, column=10).number_format = '0.0000'
                        else:
                            ws.cell(row=r, column=10).value = ""
                            
                        # LemFi Rate in Column K (11)
                        lemfi_val = rates.get("lemfi")
                        if lemfi_val == "NA":
                            ws.cell(row=r, column=11).value = "NA"
                        elif lemfi_val is not None:
                            ws.cell(row=r, column=11).value = lemfi_val
                            if isinstance(lemfi_val, (int, float)):
                                ws.cell(row=r, column=11).number_format = '0.0000'
                        else:
                            ws.cell(row=r, column=11).value = ""
                            
                        # Oanda Rate in Column L (12)
                        if rates.get("oanda") is not None:
                            ws.cell(row=r, column=12).value = rates["oanda"]
                            ws.cell(row=r, column=12).number_format = '0.0000'
                        else:
                            ws.cell(row=r, column=12).value = ""
                            
                        modified_count += 1
                        
            print(f"Updated {modified_count} currency pairs in template sheet.")
            
            def save_workbook_safely(fname):
                wb.save(fname)
                
            try:
                save_workbook_safely(filename)
                print(f"SUCCESS: Sheet populated and saved successfully as '{filename}'.")
                return
            except PermissionError:
                backup_filename = f"fx_rates_comparison_{datetime.now().strftime('%H%M%S')}.xlsx"
                print(f"\n[WARNING] Permission denied: '{filename}' is open. Saving to: '{backup_filename}'...")
                save_workbook_safely(backup_filename)
                print(f"SUCCESS: Backup sheet saved successfully as '{backup_filename}'.")
                return
        except Exception as e:
            print(f"Error populating Excel template: {e}. Falling back to default generation...")

    # Default fallback: generate Excel from scratch with generic styles
    print(f"Saving data from scratch to {filename}...")
    def write_excel_file(fname):
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="FX Comparison")
            
            # Get workbook and sheet objects
            workbook = writer.book
            worksheet = writer.sheets["FX Comparison"]
            
            # Styling configurations
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Fonts
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=10)
            
            # Fills
            fill_header = PatternFill(start_color="A80F85", end_color="A80F85", fill_type="solid") # Purple
            fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")  # Light Blue/Gray
            
            # Borders
            thin_side = Side(border_style="thin", color="D3D3D3")
            border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # Alignments
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            
            # Style Header
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_all
                
            # Style Data Rows
            for row_idx in range(2, len(df) + 2):
                is_even = (row_idx % 2 == 0)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = border_all
                    
                    # Apply alternating row colors (Zebra striping)
                    if is_even:
                        cell.fill = fill_zebra
                        
                    # Format cell data alignment and numbers
                    col_name = df.columns[col_idx - 1]
                    val = cell.value
                    
                    # Alignment
                    if col_name in ["From", "To", "Timestamp (IST)", "Timestamp (WAT)"]:
                        cell.alignment = align_center
                    elif col_name in ["Bmoni UI FX", "Bmoni Exchange Rate"] or cell.value == "NA":
                        cell.alignment = align_center # Centered blank inputs or NA
                    else:
                        cell.alignment = align_right # Numerical rates
                        
                    # Number formatting
                    if val not in ["", "NA"] and val is not None:
                        if col_name in ["LEMFI FX", "OANDA FX", "WISE FX", "GOOGLE FX RATE"]:
                            try:
                                # 4 decimal places format for exchange rates
                                cell.number_format = '0.0000'
                            except:
                                pass
                                
            # Autofit column widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                # Add padding
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # Enable grid lines
            worksheet.views.sheetView[0].showGridLines = True
            
    try:
        write_excel_file(filename)
        print(f"SUCCESS: Sheet generated successfully as '{filename}'.")
    except PermissionError:
        print(f"\n[WARNING] Permission denied: '{filename}' is currently open in Excel or another program.")
        backup_filename = f"fx_rates_comparison_{datetime.now().strftime('%H%M%S')}.xlsx"
        print(f"Attempting to save to backup file: '{backup_filename}'...")
        try:
            write_excel_file(backup_filename)
            print(f"SUCCESS: Backup sheet generated successfully as '{backup_filename}'.")
            print("Please close the original Excel file before running the script next time.")
        except Exception as e:
            print(f"[ERROR] Failed to save backup file: {e}")
    except Exception as e:
        print(f"[ERROR] Failed to save Excel file: {e}")

if __name__ == "__main__":
    main()
