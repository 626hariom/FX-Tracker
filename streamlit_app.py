# /// script
# dependencies = [
#   "requests",
#   "beautifulsoup4",
#   "pandas",
#   "openpyxl",
#   "streamlit",
# ]
# ///
import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from datetime import datetime, timezone, timedelta
import io
import random
import os

# Page Configuration
st.set_page_config(
    page_title="FX Rate Tracker & Excel Generator",
    page_icon="📊",
    layout="centered"
)

# Hide Streamlit elements (header, footer, menu)
hide_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_style, unsafe_allow_html=True)

# Scraping Functions
def fetch_google_rate(base, target, headers):
    pair_str = f"{base}-{target}"
    # Use random cache buster parameter to prevent CDN and intermediate caching
    url = f"https://www.google.com/finance/quote/{pair_str}?cb={random.random()}"
    try:
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            rate_elem = soup.find(class_="N6SYTe")
            if rate_elem:
                text_val = rate_elem.text.strip().replace(",", "")
                return float(text_val)
    except Exception as e:
        pass
    return None

def fetch_wise_rate(base, target, headers):
    url = "https://wise.com/rates/history+live"
    # Note: Adding custom cb parameters causes Cloudflare 403 blocks. We rely on Cache-Control headers instead.
    params = {
        "source": base,
        "target": target,
        "length": "1",
        "resolution": "hourly"
    }
    try:
        response = requests.get(url, params=params, headers=headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list) and len(data) > 0:
                return float(data[-1]["value"])
    except Exception as e:
        pass
    return None

def fetch_oanda_rate(base, target, headers):
    from datetime import datetime as _dt, timedelta as _td
    yesterday = (_dt.now(timezone.utc) - _td(days=1)).strftime("%Y-%m-%d")
    today = _dt.now(timezone.utc).strftime("%Y-%m-%d")
    try:
        response = requests.get(
            "https://fxds-public-exchange-rates-api.oanda.com/cc-api/currencies",
            params={
                "base": base,
                "quote": target,
                "data_type": "general_currency_pair",
                "start_date": yesterday,
                "end_date": today
            },
            headers={**headers, "Referer": "https://www.oanda.com/currency-converter/en/"},
            timeout=5
        )
        if response.status_code == 200:
            data = response.json()
            resp_list = data.get("response", [])
            if resp_list:
                latest = resp_list[-1]
                bid = float(latest["average_bid"])
                ask = float(latest["average_ask"])
                return round((bid + ask) / 2, 6)
    except Exception as e:
        pass
    return None

LEMFI_COUNTRIES = {
    "USD": "United States",
    "EUR": "Ireland",
    "GBP": "United Kingdom",
    "CAD": "Canada",
    "NGN": "Nigeria",
    "MXN": "Mexico",
}

def fetch_lemfi_rate(base, target, headers):
    import re as _re
    country = LEMFI_COUNTRIES.get(base, "United States")
    try:
        response = requests.post(
            "https://www.lemfi.com/api/lemonade/v2/exchange",
            json={"from": base, "to": target, "sender_country": country},
            headers={"Content-Type": "application/json", "x-app-locale": "en-gb", **headers},
            timeout=5
        )
        if response.status_code == 200:
            data = response.json().get("data", {})
            raw_rate = float(data.get("rate", 0))
            id_str = data.get("ID", "")
            digits_only = _re.sub(r'\D', '', id_str)
            divisor = int(digits_only) if digits_only else 1
            return round(raw_rate / divisor, 6)
    except Exception as e:
        pass
    return None

# App UI
st.title("📊 FX Rate Tracker & Comparison")
st.markdown("""
This web application fetches real-time FX rates across all **30 combinations** of **NGN, USD, MXN, EUR, GBP, CAD** 
and generates a styled Excel sheet with comparisons from:
* **Google Finance**
* **Wise**
* **Oanda** (referenced mid-market)
* **Lemfi** (NGN Remittance Corridors)

Columns for **Bmoni UI FX** and **Bmoni Exchange Rate** are left blank for manual inputs.
""")

# Sidebar Overrides
st.sidebar.header("⚙️ Manual Overrides (Optional)")
st.sidebar.markdown("""
Force specific baseline rates to match your screen exactly. 
If left blank, live rates will be used.
""")
lemfi_override_str = st.sidebar.text_input("LemFi USD-NGN Rate", placeholder="e.g. 1782").strip()
wise_override_str = st.sidebar.text_input("Wise USD-NGN Rate", placeholder="e.g. 1783").strip()

try:
    lemfi_override = float(lemfi_override_str) if lemfi_override_str else None
except ValueError:
    st.sidebar.error("Invalid number format for LemFi override.")
    lemfi_override = None

try:
    wise_override = float(wise_override_str) if wise_override_str else None
except ValueError:
    st.sidebar.error("Invalid number format for Wise override.")
    wise_override = None

st.sidebar.markdown("---")
st.sidebar.subheader("📄 Template Excel Sheet")
st.sidebar.markdown("Provide a template file to keep its layout and styles. Otherwise, the app uses `10072026_BMONI_fx_rate_30_pairs.xlsx` or defaults.")
uploaded_template = st.sidebar.file_uploader("Upload Excel Template (.xlsx)", type=["xlsx"])


if st.button("🚀 Generate Excel FX Sheet", type="primary"):
    currencies = ["NGN", "USD", "MXN", "EUR", "GBP", "CAD"]
    
    # Generate pairs
    pairs = []
    for base in currencies:
        for target in currencies:
            if base != target:
                pairs.append((base, target))
                
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache",
        "Expires": "0"
    }
    
    # First pass: Fetch all raw data from live APIs
    fetched_data = []
    google_rates = {} # Map of (base, target) -> rate
    
    utc_now = datetime.now(timezone.utc)
    ist_time = utc_now + timedelta(hours=5, minutes=30)
    wat_time = utc_now + timedelta(hours=1)
    ist_str = ist_time.strftime("%Y-%m-%d %H:%M:%S")
    wat_str = wat_time.strftime("%Y-%m-%d %H:%M:%S")
    
    # Progress Bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, (base, target) in enumerate(pairs, 1):
        status_text.text(f"Fetching raw data: {base} to {target} ({idx}/30)...")
        progress_bar.progress(idx / 30)
        
        google_rate = fetch_google_rate(base, target, headers)
        wise_rate = fetch_wise_rate(base, target, headers)
        oanda_rate = fetch_oanda_rate(base, target, headers)
        lemfi_rate = fetch_lemfi_rate(base, target, headers)
        
        # Fallbacks for raw data robustness
        if google_rate is None and wise_rate is not None:
            google_rate = wise_rate
        elif wise_rate is None and google_rate is not None:
            wise_rate = google_rate
            
        if google_rate is not None:
            google_rates[(base, target)] = google_rate
            
        fetched_data.append({
            "base": base,
            "target": target,
            "google_rate": google_rate,
            "wise_rate": wise_rate,
            "oanda_rate": oanda_rate,
            "lemfi_rate": lemfi_rate
        })
        time.sleep(0.1) # Shorter sleep for cloud run
        
    status_text.text("Processing custom overrides and calculating tables...")
    
    # Second pass: Process overrides and calculate final rates
    rows = []
    for item in fetched_data:
        base = item["base"]
        target = item["target"]
        google_rate = item["google_rate"]
        wise_rate = item["wise_rate"]
        oanda_rate = item["oanda_rate"]
        lemfi_rate = item["lemfi_rate"]
        
        # 1. Use real OANDA rate (scraped from OANDA public API)
        # oanda_rate is already fetched - keep as is (None if not available)
            
        # 2. Determine Wise rate (incorporating custom override if specified for NGN corridors)
        final_wise_rate = wise_rate
        if wise_override is not None:
            if target == "NGN":
                if base == "USD":
                    final_wise_rate = wise_override
                elif base in ["CAD", "GBP", "EUR"]:
                    rate_to_usd = google_rates.get((base, "USD"))
                    if rate_to_usd is not None:
                        final_wise_rate = round(rate_to_usd * wise_override, 4)
            elif base == "NGN":
                if target in ["USD", "CAD", "GBP", "EUR"]:
                    rate_usd_to_target = google_rates.get(("USD", target))
                    if rate_usd_to_target is not None:
                        final_wise_rate = round((1.0 / wise_override) * rate_usd_to_target, 6)
                        
        # 3. LemFi rate (scraped from LemFi API)
        # If override provided, use it for USD-NGN corridor
        if lemfi_override is not None and base == "USD" and target == "NGN":
            lemfi_rate = round(lemfi_override, 4)
        # lemfi_rate is already fetched - keep as is (None if not available)
        # LemFi API returns None for unsupported pairs (e.g. MXN corridors)
        if lemfi_rate is None:
            lemfi_rate = "NA"
        else:
            lemfi_rate = round(lemfi_rate, 4)
                    
        row = {
            "From": base,
            "To": target,
            "Bmoni UI FX": "",
            "Bmoni Exchange Rate": "",
            "LEMFI FX": lemfi_rate,
            "OANDA FX": oanda_rate,
            "WISE FX": final_wise_rate,
            "GOOGLE FX RATE": google_rate,
            "Timestamp (IST)": ist_str,
            "Timestamp (WAT)": wat_str
        }
        rows.append(row)
        
    status_text.success("Rate collection complete!")
    progress_bar.empty()
    
    df = pd.DataFrame(rows)
    
    # Show preview in the app
    st.subheader("📋 Rates Preview")
    # Convert to string to avoid PyArrow type serialization errors on mixed types in preview
    st.dataframe(df.astype(str))
    
    # Generate Excel in memory
    excel_buffer = io.BytesIO()
    template_loaded = False
    
    import openpyxl
    from openpyxl.styles import Font
    
    wb = None
    # 1. Try uploaded template first
    if uploaded_template is not None:
        try:
            # openpyxl can load from a BytesIO file-like object directly
            wb = openpyxl.load_workbook(uploaded_template)
            template_loaded = True
            st.success("Uploaded Excel template loaded successfully!")
        except Exception as e:
            st.error(f"Failed to load uploaded Excel template: {e}. Trying local file...")
            
    # 2. Try local template file next
    if wb is None:
        local_template_path = "10072026_BMONI_fx_rate_30_pairs.xlsx"
        if os.path.exists(local_template_path):
            try:
                wb = openpyxl.load_workbook(local_template_path)
                template_loaded = True
                st.info("Using local Excel template: `10072026_BMONI_fx_rate_30_pairs.xlsx`")
            except Exception as e:
                st.warning(f"Failed to load local template: {e}. Generating sheet from scratch...")
        else:
            st.warning("⚠️ Template file `10072026_BMONI_fx_rate_30_pairs.xlsx` not found. Generating a basic sheet from scratch (will lack custom headers, notes, and highlights). Please upload the template in the sidebar or commit it to your GitHub repository.")
                
    # 3. If template loaded, populate the latest rates
    template_success = False
    if template_loaded and wb is not None:
        try:
            # Find the active sheet or "FX Comparison"
            sheet_name = "FX Comparison" if "FX Comparison" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            
            # Write execution timestamp to cell F5
            ws.cell(row=5, column=6).value = f"Last Checked: {ist_str} (IST) / {wat_str} (WAT)"
            ws.cell(row=5, column=6).font = Font(name="Segoe UI", size=9, italic=True, color="555555")
            
            # Preserve empty cell F8 that openpyxl may corrupt on save
            if ws.cell(row=8, column=6).value is None:
                ws.cell(row=8, column=6).value = ""
            
            # Unify all headers in row 9 (columns F to L) to match the BMONI purple styling
            from openpyxl.styles import PatternFill, Alignment
            purple_fill = PatternFill(start_color="A80F85", end_color="A80F85", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for c in range(6, 13):
                cell = ws.cell(row=9, column=c)
                cell.fill = purple_fill
                cell.font = header_font
                cell.alignment = align_center
            
            # Map rates for quick lookup
            rates_map = {}
            for r in rows:
                rates_map[(r["From"], r["To"])] = {
                    "google": r["GOOGLE FX RATE"],
                    "wise": r["WISE FX"],
                    "lemfi": r["LEMFI FX"],
                    "oanda": r["OANDA FX"]
                }
                
            modified_count = 0
            # Populate cells (Rows 10 to 39, column F to L)
            for r in range(10, 40):
                pair_cell = ws.cell(row=r, column=6).value # Column F is 6
                if not pair_cell:
                    continue
                # Clean pair (replace arrows and spaces)
                cleaned = str(pair_cell).replace("→", "").replace("->", "").replace("to", "").strip()
                parts = [p.strip() for p in cleaned.split() if p.strip()]
                if len(parts) == 2:
                    base, target = parts[0], parts[1]
                    key = (base, target)
                    if key in rates_map:
                        rates = rates_map[key]
                        
                        # Google Rate in Column I (9)
                        if rates.get("google") is not None:
                            ws.cell(row=r, column=9).value = rates["google"]
                        else:
                            ws.cell(row=r, column=9).value = ""
                            
                        # Wise Rate in Column J (10)
                        if rates.get("wise") is not None:
                            ws.cell(row=r, column=10).value = rates["wise"]
                        else:
                            ws.cell(row=r, column=10).value = ""
                            
                        # LemFi Rate in Column K (11)
                        lemfi_val = rates.get("lemfi")
                        if lemfi_val == "NA":
                            ws.cell(row=r, column=11).value = "NA"
                        elif lemfi_val is not None:
                            ws.cell(row=r, column=11).value = lemfi_val
                        else:
                            ws.cell(row=r, column=11).value = ""
                            
                        # Oanda Rate in Column L (12)
                        if rates.get("oanda") is not None:
                            ws.cell(row=r, column=12).value = rates["oanda"]
                        else:
                            ws.cell(row=r, column=12).value = ""
                            
                        modified_count += 1
                        
            # Save the workbook to our memory buffer
            wb.save(excel_buffer)
            template_success = True
            st.success(f"Populated {modified_count} currency pairs inside your designed Excel template!")
        except Exception as e:
            st.error(f"Error filling template: {e}. Falling back to default styled table...")
            template_success = False
            
    # 4. Fallback: generate default styled table from scratch
    if not template_success:
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="FX Comparison")
            
            # Styling configurations
            workbook = writer.book
            worksheet = writer.sheets["FX Comparison"]
            
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=10)
            
            fill_header = PatternFill(start_color="A80F85", end_color="A80F85", fill_type="solid")
            fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            
            thin_side = Side(border_style="thin", color="D3D3D3")
            border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_all
                
            for row_idx in range(2, len(df) + 2):
                is_even = (row_idx % 2 == 0)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = border_all
                    
                    if is_even:
                        cell.fill = fill_zebra
                        
                    col_name = df.columns[col_idx - 1]
                    val = cell.value
                    
                    if col_name in ["From", "To", "Timestamp (IST)", "Timestamp (WAT)"]:
                        cell.alignment = align_center
                    elif col_name in ["Bmoni UI FX", "Bmoni Exchange Rate"] or cell.value == "NA":
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        
                    if val not in ["", "NA"] and val is not None:
                        if col_name in ["LEMFI FX", "OANDA FX", "WISE FX", "GOOGLE FX RATE"]:
                            try:
                                cell.number_format = '0.0000'
                            except:
                                pass
                                
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            worksheet.views.sheetView[0].showGridLines = True
            
    excel_buffer.seek(0)
    
    # Download Button
    st.download_button(
        label="📥 Download Excel Sheet",
        data=excel_buffer,
        file_name=f"fx_rates_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
